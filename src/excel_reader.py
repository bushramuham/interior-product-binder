"""Read an exported schedule .xlsx and group rows into Products by shared PATH."""

import os
from urllib.parse import urlsplit

from openpyxl import Workbook, load_workbook

from src.models import Product, ProductEntry

# Canonical schedule columns (order used when writing a new schedule).
SCHEDULE_HEADERS = ["KEY", "DESCRIPTION", "PG", "PATH", "DESCRIPTION2", "DESCRIPTION3"]

# Recognized header names (lowercased) -> canonical field
_HEADER_ALIASES = {
    "key": "key",
    "description": "description",
    "description2": "description2",
    "description3": "description3",
    "pg": "page_hint",
    "page": "page_hint",
    "path": "raw_path",
    "source": "raw_path",
}


def _normalize_source(path: str) -> tuple[str, str]:
    """Return (group_id, source_type) for a PATH cell value."""
    path = path.strip()
    lower = path.lower()
    if lower.startswith(("http://", "https://")):
        parts = urlsplit(path)
        group_id = f"{parts.scheme.lower()}://{parts.netloc.lower()}{parts.path}"
        if parts.query:
            group_id += f"?{parts.query}"
        return group_id, "url"
    return os.path.normcase(os.path.normpath(path)), "pdf"


def save_schedule(rows: list[dict], path: str) -> None:
    """Write rows (dicts keyed by SCHEDULE_HEADERS) to an .xlsx schedule."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MC Schedule by Keynote"
    ws.append(SCHEDULE_HEADERS)
    for row in rows:
        ws.append([str(row.get(h, "") or "") for h in SCHEDULE_HEADERS])
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb.save(path)


# Canonical field name -> schedule header (for load_rows round-tripping).
_FIELD_TO_HEADER = {
    "key": "KEY", "description": "DESCRIPTION", "page_hint": "PG",
    "raw_path": "PATH", "description2": "DESCRIPTION2", "description3": "DESCRIPTION3",
}


def load_rows(xlsx_path: str) -> list[dict]:
    """Read a schedule into flat row dicts keyed by SCHEDULE_HEADERS (for
    review/editing in the front-ends). Relative PATH values are made absolute
    against the spreadsheet's own folder when they resolve there, so the rows
    can be re-saved anywhere without breaking."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        wb.close()
        raise ValueError(f"Spreadsheet is empty: {xlsx_path}")
    columns = _map_headers(header_row)
    if "key" not in columns.values():
        wb.close()
        raise ValueError(
            f"Could not find the required KEY column in {xlsx_path}. "
            f"Found headers: {[str(c) for c in header_row if c is not None]}"
        )

    xlsx_dir = os.path.dirname(os.path.abspath(xlsx_path))
    out: list[dict] = []
    for row in it:
        rec = {h: "" for h in SCHEDULE_HEADERS}
        for idx, fieldname in columns.items():
            cell = row[idx] if idx < len(row) else None
            rec[_FIELD_TO_HEADER[fieldname]] = str(cell).strip() if cell is not None else ""
        if not rec["KEY"]:
            continue
        path = rec["PATH"]
        if path and not path.lower().startswith(("http://", "https://")) \
                and not os.path.isabs(path):
            if os.path.exists(path):                     # CWD-relative
                rec["PATH"] = os.path.abspath(path)
            else:                                        # schedule-relative
                candidate = os.path.normpath(os.path.join(xlsx_dir, path))
                if os.path.exists(candidate):
                    rec["PATH"] = candidate
        out.append(rec)
    wb.close()
    return out


def _map_headers(header_row) -> dict[int, str]:
    """Map column index -> canonical field name, case-insensitively."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        if name in _HEADER_ALIASES:
            mapping[idx] = _HEADER_ALIASES[name]
    return mapping


def load_products(xlsx_path: str) -> list[Product]:
    """Load the schedule and return Products grouped by normalized PATH, in row order."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        wb.close()
        raise ValueError(f"Spreadsheet is empty: {xlsx_path}")

    columns = _map_headers(header_row)
    if "key" not in columns.values():
        wb.close()
        raise ValueError(
            f"Could not find the required KEY column in {xlsx_path}. "
            f"Found headers: {[str(c) for c in header_row if c is not None]}"
        )

    groups: dict[str, Product] = {}
    for row_num, row in enumerate(rows, start=2):
        values = {}
        for idx, fieldname in columns.items():
            cell = row[idx] if idx < len(row) else None
            values[fieldname] = str(cell).strip() if cell is not None else ""

        entry = ProductEntry(
            key=values.get("key", ""),
            description=values.get("description", ""),
            description2=values.get("description2", ""),
            description3=values.get("description3", ""),
            raw_path=values.get("raw_path", ""),
            page_hint=values.get("page_hint", ""),
        )

        if not entry.key:
            if any([entry.description, entry.raw_path]):
                print(f"  WARNING: skipping row {row_num} (missing KEY)")
            continue

        if entry.raw_path:
            group_id, source_type = _normalize_source(entry.raw_path)
        elif entry.page_hint:
            # Sourceless rows that share a PG value are grouped onto one
            # OWNER INPUT NEEDED page.
            group_id, source_type = f"__nosource__pg_{entry.page_hint.lower()}", "none"
        else:
            # Sourceless row with no PG -> its own OWNER INPUT NEEDED page.
            group_id, source_type = f"__nosource__row_{row_num}", "none"

        if group_id not in groups:
            groups[group_id] = Product(
                group_id=group_id,
                source_path=entry.raw_path.strip(),
                source_type=source_type,
            )
        groups[group_id].entries.append(entry)

    wb.close()
    # Lexicographic order everywhere: keys within each group, then groups by
    # their first key. This drives page order and the TOC.
    products = list(groups.values())
    for p in products:
        p.entries.sort(key=lambda e: e.key.upper())
    products.sort(key=lambda p: p.entries[0].key.upper())
    return products
